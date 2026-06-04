#!/usr/bin/env python3
"""
log-metrics.py — Atualiza as Métricas no dashboard Excel a partir do step-log.ndjson.

Uso:
  python3 log-metrics.py                    # Mostra métricas no terminal
  python3 log-metrics.py --update-excel      # Atualiza o dashboard Excel
  python3 log-metrics.py --days 14           # Últimos 14 dias (default: 30)
  python3 log-metrics.py --watch             # Mostra em tempo real (tail -f style)
"""

import json
import os
import sys
from datetime import datetime, timedelta, timezone

LOG_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "logs")
STEP_LOG = os.path.join(LOG_DIR, "step-log.ndjson")
DASHBOARD_FILE = os.path.expanduser("~/Development/teamwill/mobilize/workflow/karpathy/wiki/references/ai-skills-dashboard.xlsx")


def read_log():
    entries = []
    if not os.path.exists(STEP_LOG):
        print(f"\033[90mFicheiro de log não encontrado: {STEP_LOG}\033[0m")
        return entries
    with open(STEP_LOG) as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    entries.append(json.loads(line))
                except json.JSONDecodeError:
                    pass
    return entries


def compute_metrics(entries, days=30):
    cutoff = datetime.now(timezone.utc) - timedelta(days=days) if days else datetime.min.replace(tzinfo=timezone.utc)
    recent = [e for e in entries if datetime.fromisoformat(e.get("timestamp", "")).replace(tzinfo=timezone.utc) >= cutoff]

    workflows = {}
    for e in recent:
        wf = e.get("workflow_id")
        if wf:
            if wf not in workflows:
                workflows[wf] = {"steps": [], "start_time": None, "end_time": None}
            workflows[wf]["steps"].append(e)
            if e.get("event") == "start":
                workflows[wf]["start_time"] = e.get("timestamp")
            if e.get("event") == "end":
                workflows[wf]["end_time"] = e.get("timestamp")

    # Per-agent stats
    agent_stats = {}
    for e in recent:
        if e.get("event") != "end":
            continue
        agent = e.get("agent", "unknown")
        if agent not in agent_stats:
            agent_stats[agent] = {"total": 0, "success": 0, "failure": 0, "skipped": 0, "durations": []}
        agent_stats[agent]["total"] += 1
        status = e.get("status", "")
        if status == "success":
            agent_stats[agent]["success"] += 1
        elif status == "failure":
            agent_stats[agent]["failure"] += 1
        elif status == "skipped":
            agent_stats[agent]["skipped"] += 1
        dur = e.get("duration_ms")
        if dur:
            agent_stats[agent]["durations"].append(dur)

    # Failures
    failures = [e for e in recent if e.get("event") == "end" and e.get("status") == "failure"]

    # Workflow-level stats
    wf_summaries = []
    for wf_id, data in workflows.items():
        ends = [s for s in data["steps"] if s.get("event") == "end"]
        successes = sum(1 for s in ends if s.get("status") == "success")
        failures_wf = sum(1 for s in ends if s.get("status") == "failure")
        total_dur = 0
        if data.get("start_time") and data.get("end_time"):
            try:
                start = datetime.fromisoformat(data["start_time"])
                end = datetime.fromisoformat(data["end_time"])
                total_dur = (end - start).total_seconds()
            except Exception:
                pass
        wf_summaries.append({
            "id": wf_id,
            "date": data["start_time"][:10] if data.get("start_time") else "?",
            "duration_s": round(total_dur, 1),
            "steps_total": len(ends),
            "successes": successes,
            "failures": failures_wf,
            "status": "✅" if failures_wf == 0 else "❌",
        })

    wf_summaries.sort(key=lambda x: x["date"], reverse=True)

    return {
        "timeframe_days": days,
        "workflows": wf_summaries,
        "agents": agent_stats,
        "failures": sorted(failures, key=lambda x: x.get("timestamp", ""), reverse=True),
        "total_entries": len(recent),
    }


def print_metrics(metrics):
    print(f"\n{'═' * 60}")
    print(f"  📊 Métricas Operacionais (últimos {metrics['timeframe_days']} dias)")
    print(f"{'═' * 60}")

    # Workflow summaries
    if metrics["workflows"]:
        print(f"\n  {'Workflow':16} {'Data':12} {'Dur(s)':8} {'Steps':6} {'✅':6} {'❌':6} {'Estado':8}")
        print(f"  {'─'*62}")
        for w in metrics["workflows"]:
            print(f"  {w['id']:16} {w['date']:12} {w['duration_s']:>8} {w['steps_total']:>6} {w['successes']:>6} {w['failures']:>6} {w['status']:>8}")
    else:
        print(f"\n  \033[90mNenhum workflow nos últimos {metrics['timeframe_days']} dias.\033[0m")

    # Per-agent
    print(f"\n  {'Agente':22} {'Steps':6} {'✅':6} {'❌':6} {'⏭️':6} {'Taxa':8} {'Média (s)':10}")
    print(f"  {'─'*64}")
    for agent in sorted(metrics["agents"].keys()):
        s = metrics["agents"][agent]
        rate = round(s["success"] / s["total"] * 100, 1) if s["total"] else 0
        avg_dur = round(sum(s["durations"]) / len(s["durations"]) / 1000, 1) if s["durations"] else 0
        rate_color = "\033[32m" if rate >= 80 else "\033[33m" if rate >= 50 else "\033[31m"
        print(f"  \033[36m{agent:22}\033[0m {s['total']:>6} {s['success']:>6} {s['failure']:>6} {s['skipped']:>6} {rate_color}{rate:>6.1f}%\033[0m {avg_dur:>8.1f}s")

    # Failures
    if metrics["failures"]:
        print(f"\n  \033[31m❌ Últimas falhas:\033[0m")
        for f in metrics["failures"][:5]:
            ts = f.get("timestamp", "")[11:19]
            print(f"    [{ts}] \033[35m{f.get('workflow_id', '?')}\033[0m / \033[36m{f.get('agent', '?')}\033[0m / \033[94m{f.get('step', '?')}\033[0m — \033[90m{f.get('error', 'sem detalhe')[:80]}\033[0m")

    print()


def update_excel(metrics):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl não instalado. Instala com: pip install openpyxl")
        return

    if not os.path.exists(DASHBOARD_FILE):
        print(f"Dashboard não encontrado: {DASHBOARD_FILE}")
        return

    wb = openpyxl.load_workbook(DASHBOARD_FILE)
    THIN_BORDER = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )
    BODY_FONT = Font(name="Inter", size=10)

    # Sheet 3: Métricas
    if "Métricas" not in wb.sheetnames:
        print("Sheet 'Métricas' não encontrada no dashboard.")
        return

    ws = wb["Métricas"]

    # Clear old data (rows 6+)
    for row in range(6, 50):
        for col in range(1, 8):
            ws.cell(row=row, column=col, value=None)

    # Populate per-workflow (row 6+)
    for i, w in enumerate(metrics["workflows"]):
        row = 6 + i
        vals = [w["id"], w["date"], w["duration_s"], w["steps_total"], w["successes"], w["failures"], w["status"]]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center")
            cell.border = THIN_BORDER

    # Populate per-agent (starting at row 9+1=10)
    for i, agent in enumerate(sorted(metrics["agents"].keys())):
        row = 10 + i
        s = metrics["agents"][agent]
        rate = round(s["success"] / s["total"] * 100, 1) if s["total"] else 0
        avg_dur = round(sum(s["durations"]) / len(s["durations"]) / 1000, 1) if s["durations"] else 0
        vals = [agent, "—", s["total"], s["success"], rate, avg_dur]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center")
            cell.border = THIN_BORDER

    # Populate failures
    fail_start = 19
    if metrics["failures"]:
        for i, f in enumerate(metrics["failures"][:10]):
            row = fail_start + 2 + i
            ts = f.get("timestamp", "")[:19]
            vals = [ts, f.get("workflow_id", ""), f.get("step", ""), f.get("agent", ""),
                    f.get("duration_ms", ""), f.get("error", "")[:60] if f.get("error") else ""]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = BODY_FONT
                cell.alignment = Alignment(vertical="center")
                cell.border = THIN_BORDER
    else:
        ws.cell(row=fail_start + 2, column=1, value="(sem falhas registadas)")

    wb.save(DASHBOARD_FILE)
    print(f"✅ Dashboard atualizado: {DASHBOARD_FILE}")


if __name__ == "__main__":
    days = 30
    update = False
    watch = False

    args = sys.argv[1:]
    i = 0
    while i < len(args):
        if args[i] == "--days":
            i += 1
            days = int(args[i])
        elif args[i] == "--update-excel":
            update = True
        elif args[i] == "--watch":
            watch = True
        i += 1

    if watch:
        import time
        try:
            while True:
                os.system("clear")
                entries = read_log()
                metrics = compute_metrics(entries, days)
                print_metrics(metrics)
                time.sleep(5)
        except KeyboardInterrupt:
            print("\n👋")
    else:
        entries = read_log()
        if not entries:
            print("\033[90mNenhum log encontrado no ficheiro step-log.ndjson.\033[0m")
            sys.exit(0)
        metrics = compute_metrics(entries, days)
        print_metrics(metrics)
        if update:
            update_excel(metrics)
