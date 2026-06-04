#!/usr/bin/env python3
"""
prompt-register.py — Regista versões de prompts dos agentes/skills no dashboard Excel.

Uso:
  python3 prompt-register.py scan              # Mostra versões actuais vs dashboard
  python3 prompt-register.py scan --excel      # Atualiza dashboard com novas versões
  python3 prompt-register.py add <skill> <versão> <cat> <desc> <resultado> <nota> [notas]
  python3 prompt-register.py log <workflow_id>  # Regista alteração no step-log
"""

import json
import os
import re
import subprocess
import sys
from datetime import date

WORKFLOW = os.path.expanduser("~/Development/teamwill/mobilize/workflow")
CONFIG_DIR = os.path.expanduser("~/.config/opencode")
DASHBOARD = os.path.join(WORKFLOW, "karpathy", "wiki", "references", "ai-skills-dashboard.xlsx")
STEP_LOG = os.path.join(WORKFLOW, "logs", "step-log.ndjson")

AGENTS_DIR = os.path.join(CONFIG_DIR, "agents")
SKILLS_DIR = os.path.join(CONFIG_DIR, "skills")


def extract_frontmatter(path):
    """Extract YAML frontmatter fields from agent/skill .md file."""
    if not os.path.isfile(path):
        return {}
    with open(path) as f:
        content = f.read()
    # Match --- ... ---
    m = re.match(r'^---\s*\n(.*?)\n---', content, re.DOTALL)
    if not m:
        return {}
    data = {}
    for line in m.group(1).split('\n'):
        if ':' in line:
            key, _, val = line.partition(':')
            data[key.strip()] = val.strip()
    # Normalize version: strip leading "v" if present
    if "version" in data:
        data["version"] = data["version"].lstrip("v")
    return data


def scan_agents():
    """Scan all agent .md files for version info."""
    results = []
    if not os.path.isdir(AGENTS_DIR):
        return results
    for fname in sorted(os.listdir(AGENTS_DIR)):
        if not fname.endswith('.md'):
            continue
        path = os.path.join(AGENTS_DIR, fname)
        fm = extract_frontmatter(path)
        results.append({
            "name": fm.get("name", fname.replace('.md', '')),
            "version": fm.get("version", "?"),
            "model": fm.get("model", "?"),
            "file": fname,
        })
    return results


def scan_skills():
    """Scan all skill files for version info."""
    results = []
    if not os.path.isdir(SKILLS_DIR):
        return results
    for name in sorted(os.listdir(SKILLS_DIR)):
        # Check both dir/SKILL.md and name.md
        paths = [
            os.path.join(SKILLS_DIR, name, "SKILL.md"),
            os.path.join(SKILLS_DIR, f"{name}.md"),
        ]
        found = None
        for p in paths:
            if os.path.isfile(p):
                found = p
                break
        if not found:
            continue
        fm = extract_frontmatter(found)
        results.append({
            "name": fm.get("name", name),
            "version": fm.get("version", "?"),
            "file": os.path.basename(found),
        })
    return results


def read_dashboard_versions():
    """Read existing versions from the dashboard Excel."""
    try:
        import openpyxl
    except ImportError:
        return set()

    if not os.path.isfile(DASHBOARD):
        return set()

    wb = openpyxl.load_workbook(DASHBOARD)
    if "Registo" not in wb.sheetnames:
        return set()

    ws = wb["Registo"]
    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            existing.add((str(row[0]).strip(), str(row[1]).strip()))
    wb.close()
    return existing


def cmd_scan(update_excel=False):
    agents = scan_agents()
    skills = scan_skills()
    existing = read_dashboard_versions()

    print(f"\n{'═' * 60}")
    print(f"  📋 Estado dos Prompts")
    print(f"{'═' * 60}")
    print(f"\n  \033[1mAgentes\033[0m")
    for a in agents:
        key = (a["name"], a["version"])
        in_dash = "✅" if key in existing else "⬜"
        model_short = a["model"].split("/")[-1][:20] if "/" in a["model"] else a["model"][:20]
        print(f"  {in_dash} \033[36m{a['name']:24}\033[0m v{a['version']:8} {model_short}")

    print(f"\n  \033[1mSkills\033[0m")
    for s in skills:
        key = (s["name"], s["version"])
        in_dash = "✅" if key in existing else "⬜"
        print(f"  {in_dash} \033[36m{s['name']:28}\033[0m v{s['version']:8}")

    # Check for missing entries
    missing = [(a["name"], a["version"]) for a in agents if (a["name"], a["version"]) not in existing]
    missing += [(s["name"], s["version"]) for s in skills if (s["name"], s["version"]) not in existing]

    if missing:
        print(f"\n  \033[33m⚠ {len(missing)} versões não registadas no dashboard:\033[0m")
        for name, ver in missing:
            print(f"    → \033[36m{name}\033[0m v{ver}")
        if not update_excel:
            print(f"\n  Para registar: python3 prompt-register.py scan --excel")
        else:
            _auto_register(missing)
    else:
        print(f"\n  \033[32m✓ Todas as versões estão registadas no dashboard.\033[0m")

    print()


def _auto_register(missing):
    """Auto-register missing versions in the dashboard Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  openpyxl não instalado.")
        return

    if not os.path.isfile(DASHBOARD):
        print(f"  Dashboard não encontrado: {DASHBOARD}")
        return

    wb = openpyxl.load_workbook(DASHBOARD)
    ws = wb["Registo"]

    # Find last row
    last_row = ws.max_row + 1

    BODY_FONT = Font(name="Inter", size=10)
    THIN_BORDER = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )

    for name, ver in missing:
        row = last_row
        last_row += 1
        data = [name, ver, date.today(), "(auto-registado)", "Outro", "Registo automático de versão", "", None, ""]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    wb.save(DASHBOARD)
    print(f"  ✅ {len(missing)} versões registadas no dashboard.")


def cmd_add(name, version, category, description, result, score, notes=""):
    """Manually add a version entry to the dashboard."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl não instalado.")
        return

    if not os.path.isfile(DASHBOARD):
        print(f"Dashboard não encontrado: {DASHBOARD}")
        return

    wb = openpyxl.load_workbook(DASHBOARD)
    ws = wb["Registo"]
    last_row = ws.max_row + 1

    BODY_FONT = Font(name="Inter", size=10)
    THIN_BORDER = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )
    SCORE_FILLS = {
        5: PatternFill(patternType="solid", fgColor="4CAF50"),
        4: PatternFill(patternType="solid", fgColor="9BDE7E"),
    }

    data = [name, version, date.today(), "", category, description, result, score, notes]
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=last_row, column=col, value=val)
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Color the score
    score_cell = ws.cell(row=last_row, column=8)
    if score in SCORE_FILLS:
        score_cell.fill = SCORE_FILLS[score]
        score_cell.font = Font(name="Inter", bold=True, size=12, color="FFFFFF")
    score_cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(DASHBOARD)
    print(f"✅ Entrada registada: {name} v{version} no dashboard.")


def cmd_log(wf_id):
    """Log prompt version event in step-log."""
    try:
        result = subprocess.run(
            ["python3", os.path.join(WORKFLOW, "scripts", "step-log.py"),
             "log", wf_id, "prompt-register", "verify", "info", "Versões de prompts verificadas"],
            capture_output=True, text=True, timeout=10
        )
        print(result.stdout.strip())
    except Exception as e:
        print(f"Erro: {e}")


def usage():
    print(__doc__)
    sys.exit(1)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        usage()
    cmd = sys.argv[1]

    if cmd == "scan":
        update_excel = "--excel" in sys.argv
        cmd_scan(update_excel)

    elif cmd == "add":
        if len(sys.argv) < 7:
            usage()
        name = sys.argv[2]
        version = sys.argv[3]
        category = sys.argv[4]
        description = sys.argv[5]
        result = sys.argv[6]
        score = int(sys.argv[7]) if len(sys.argv) > 7 and sys.argv[7].isdigit() else None
        notes = " ".join(sys.argv[8:]) if len(sys.argv) > 8 else ""
        cmd_add(name, version, category, description, result, score, notes)

    elif cmd == "log":
        if len(sys.argv) < 3:
            usage()
        cmd_log(sys.argv[2])

    else:
        usage()
