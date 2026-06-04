#!/usr/bin/env python3
import sys
import json
from openpyxl import load_workbook


def read_excel_file(file_path: str, sheet_name: str = None) -> dict:
    try:
        wb = load_workbook(file_path, data_only=True)
        
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return {"error": f"Sheet not found: {sheet_name}", "success": False}
            ws = wb[sheet_name]
            data = []
            headers = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_idx == 1:
                    headers = [cell if cell is not None else f"__EMPTY_{i}" for i, cell in enumerate(row)]
                else:
                    row_dict = {headers[i]: cell for i, cell in enumerate(row) if i < len(headers)}
                    data.append(row_dict)
            return {
                "success": True,
                "file": file_path,
                "sheet": sheet_name,
                "row_count": len(data),
                "columns": headers,
                "data": data
            }
        else:
            sheets = {}
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                data = []
                headers = []
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row_idx == 1:
                        headers = [cell if cell is not None else f"__EMPTY_{i}" for i, cell in enumerate(row)]
                    else:
                        row_dict = {headers[i]: cell for i, cell in enumerate(row) if i < len(headers)}
                        if any(row_dict.values()):
                            data.append(row_dict)
                sheets[sheet] = {
                    "row_count": len(data),
                    "columns": headers,
                    "data": data
                }
            return {"success": True, "file": file_path, "sheets": sheets}
    except FileNotFoundError:
        return {"error": f"File not found: {file_path}", "success": False}
    except Exception as e:
        return {"error": str(e), "success": False}


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 read_excel.py <file.xlsx> [sheet_name]")
        sys.exit(1)
    
    file_path = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None
    
    result = read_excel_file(file_path, sheet_name)
    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()