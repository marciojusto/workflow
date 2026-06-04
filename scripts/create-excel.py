import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule
from datetime import date, timedelta
import os

FILE = os.path.expanduser("~/Development/teamwill/mobilize/workflow/karpathy/wiki/references/ai-skills-dashboard.xlsx")

wb = openpyxl.Workbook()

# ── Color palette ──
HEADER_FILL = PatternFill(patternType="solid", fgColor="1A1A2E")
HEADER_FONT = Font(name="Inter", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Inter", size=10)
BODY_FONT_MONO = Font(name="JetBrains Mono", size=9)
SCORE_FILLS = {
    1: PatternFill(patternType="solid", fgColor="FF4C4C"),
    2: PatternFill(patternType="solid", fgColor="FF8C42"),
    3: PatternFill(patternType="solid", fgColor="FFD166"),
    4: PatternFill(patternType="solid", fgColor="9BDE7E"),
    5: PatternFill(patternType="solid", fgColor="4CAF50"),
}
THIN_BORDER = Border(
    left=Side(style="thin", color="E0E0E0"),
    right=Side(style="thin", color="E0E0E0"),
    top=Side(style="thin", color="E0E0E0"),
    bottom=Side(style="thin", color="E0E0E0"),
)
ZEBRA_ODD = PatternFill(patternType="solid", fgColor="F5F5FA")
ZEBRA_EVEN = PatternFill(patternType="solid", fgColor="FFFFFF")

# ════════════════════════════════════════════
# SHEET 1 — LOG (raw entries)
# ════════════════════════════════════════════
ws = wb.active
ws.title = "Registo"

HEADERS = [
    ("Skill", 22),
    ("Versão", 10),
    ("Data Versão", 14),
    ("Prompt Exato (abreviado)", 60),
    ("Categoria Alteração", 20),
    ("Descrição Curta", 40),
    ("Resultado Observado", 50),
    ("Performance\n(1-5)", 12),
    ("Notas", 30),
]

for col, (hdr, width) in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=hdr)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(col)].width = width

ws.row_dimensions[1].height = 36

# Freeze header
ws.freeze_panes = "A2"

# Auto filter
ws.auto_filter.ref = f"A1:I{ws.max_row}"

# Data validation for Score (1-5)
dv_score = DataValidation(type="whole", operator="between", formula1=1, formula2=5, allow_blank=True)
dv_score.error = "A nota deve ser entre 1 e 5."
dv_score.errorTitle = "Nota inválida"
dv_score.prompt = "Insira 1-5"
dv_score.promptTitle = "Performance"
ws.add_data_validation(dv_score)

# Data validation for Category
CATEGORIES = "Mudança de Tom,Foco/Âmbito,Restrição,Modelo,Temperatura,Contexto,Persona,Formato,Estrutura,Correção Bug,Outro"
dv_cat = DataValidation(type="list", formula1=f'"{CATEGORIES}"', allow_blank=True)
dv_cat.prompt = "Seleccione categoria"
dv_cat.promptTitle = "Categoria"
ws.add_data_validation(dv_cat)

# ── Sample data ──
samples = [
    ["miles-expert", "1.0.0", date(2026, 5, 1), "You are an expert in European automotive leasing...", "Estrutura", "Versão inicial com workflow básico", "Funciona mas verboso", 3, ""],
    ["miles-expert", "1.1.0", date(2026, 5, 10), "You are an expert in European automotive... (refinado)", "Foco/Âmbito", "Adicionado step 0.6 review-plan e validação", "Fluxo mais robusto", 4, "Melhorou aprovação"],
    ["workflow-orchestrator", "1.0.0", date(2026, 5, 5), "You are a workflow orchestrator...", "Estrutura", "Criação inicial com multi-mode (AUTO/MANUAL)", "Funcional", 4, ""],
    ["wiki-keeper", "1.0.0", date(2026, 4, 28), "You manage the karpathy wiki...", "Persona", "Primeira versão com ingestão e query", "Base sólida", 4, ""],
    ["wiki-keeper", "1.0.1", date(2026, 5, 3), "Same + Atlassian MCP read-only rules", "Restrição", "Adicionada regra READ-ONLY no Atlassian MCP", "Segurança reforçada", 5, "Evita writes acidentais"],
    ["coherence-checker", "1.0.0", date(2026, 5, 12), "You validate implementation coherence...", "Persona", "Criação do agente de validação", "Útil para code review", 4, ""],
    ["review-plan", "1.0.0", date(2026, 5, 12), "You are an independent plan reviewer...", "Persona", "Criação do revisor de planos", "Bom reasoning", 5, "GLM-5.1 recomendado"],
    ["miles-expert", "1.2.0", date(2026, 5, 14), "You are an expert... (model selection v1.1.0)", "Modelo", "Estratégia de selecção de modelo: default M2.7, escalada para DeepSeek V4 Pro", "Custo-benefício ideal", 5, "Grande ganho"],
]

for row_idx, row_data in enumerate(samples, 2):
    fill = ZEBRA_ODD if row_idx % 2 == 0 else ZEBRA_EVEN
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = BODY_FONT_MONO if col_idx == 4 else BODY_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        cell.fill = fill
    # Score formatting
    score_cell = ws.cell(row=row_idx, column=8)
    score_val = row_data[7]
    if score_val and score_val in SCORE_FILLS:
        score_cell.fill = SCORE_FILLS[score_val]
        score_cell.font = Font(name="Inter", bold=True, size=12, color="FFFFFF")
    score_cell.alignment = Alignment(horizontal="center", vertical="center")
    dv_score.add(score_cell)
    # Category validation
    dv_cat.add(ws.cell(row=row_idx, column=5))

# Row height for data
for r in range(2, len(samples) + 2):
    ws.row_dimensions[r].height = 36

# Conditional formatting — DataBar on Score column
ws.conditional_formatting.add(
    f"H2:H{len(samples) + 1}",
    DataBarRule(start_type="min", end_type="max", color="4CAF50", showValue=True)
)

# ════════════════════════════════════════════
# SHEET 2 — SUMMARY
# ════════════════════════════════════════════
ws2 = wb.create_sheet("Resumo")

# Title
ws2.cell(row=1, column=1, value="📊 Dashboard de Skills IA").font = Font(name="Inter", bold=True, size=16, color="1A1A2E")
ws2.merge_cells("A1:H1")
ws2.row_dimensions[1].height = 40

# ── Table: Avg Score per Skill ──
ws2.cell(row=3, column=1, value="Média Performance por Skill").font = Font(name="Inter", bold=True, size=12)
ws2.merge_cells("A3:C3")

summary_headers = ["Skill", "Média", "Total Versões"]
for col, hdr in enumerate(summary_headers, 1):
    cell = ws2.cell(row=4, column=col, value=hdr)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

# Get unique skills from sample data
skills = sorted(set(s[0] for s in samples))
for i, skill in enumerate(skills):
    versions = [s for s in samples if s[0] == skill]
    avg = sum(s[7] for s in versions) / len(versions)
    row = 5 + i
    ws2.cell(row=row, column=1, value=skill).font = BODY_FONT
    ws2.cell(row=row, column=1).border = THIN_BORDER
    avg_cell = ws2.cell(row=row, column=2, value=round(avg, 1))
    avg_cell.font = Font(name="Inter", bold=True, size=11)
    avg_cell.alignment = Alignment(horizontal="center")
    avg_cell.border = THIN_BORDER
    ws2.cell(row=row, column=3, value=len(versions)).font = BODY_FONT
    ws2.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    ws2.cell(row=row, column=3).border = THIN_BORDER

# ── Table: Best Improvements ──
best_start = 5 + len(skills) + 2
ws2.cell(row=best_start, column=1, value="🏆 Melhores Incrementos (>4.0)").font = Font(name="Inter", bold=True, size=12)
ws2.merge_cells(f"A{best_start}:F{best_start}")

best_headers = ["Skill", "Versão", "Data", "Descrição", "Nota", "Ganho"]
for col, hdr in enumerate(best_headers, 1):
    cell = ws2.cell(row=best_start + 1, column=col, value=hdr)
    cell.font = HEADER_FONT
    cell.fill = PatternFill(patternType="solid", fgColor="2E7D32")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

best_samples = [s for s in samples if s[7] >= 4]
best_samples.sort(key=lambda x: x[7], reverse=True)
for i, s in enumerate(best_samples):
    row = best_start + 2 + i
    vals = [s[0], s[1], s[2], s[5], s[7], "✅" if s[7] >= 4 else "➖"]
    for col, val in enumerate(vals, 1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN_BORDER
    if s[7] == 5:
        ws2.cell(row=row, column=5).fill = SCORE_FILLS[5]
        ws2.cell(row=row, column=5).font = Font(name="Inter", bold=True, size=12, color="FFFFFF")
        ws2.cell(row=row, column=5).alignment = Alignment(horizontal="center")

ws2.column_dimensions["A"].width = 24
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 50
ws2.column_dimensions["E"].width = 10
ws2.column_dimensions["F"].width = 10

# ── Chart: Performance Trend ──
chart = LineChart()
chart.title = "Evolução de Performance por Versão"
chart.x_axis.title = "Versão (índice)"
chart.y_axis.title = "Nota (1-5)"
chart.y_axis.scaling.min = 1
chart.y_axis.scaling.max = 5
chart.style = 10
chart.width = 28
chart.height = 16

# Group data by skill for chart
for skill in skills:
    versions = sorted(
        [(s[2], s[7], f"{s[0]} {s[1]}") for s in samples if s[0] == skill],
        key=lambda x: x[0]
    )
    if len(versions) < 2:
        continue
    # Write chart data to temporary columns
    start_col = skills.index(skill) * 3 + 1
    start_row = best_start + 4 + len(samples) + 2
    ws2.cell(row=start_row, column=start_col, value=f"{skill} (data)").font = Font(size=8)
    ws2.cell(row=start_row, column=start_col + 1, value=f"{skill} (nota)").font = Font(size=8)
    for i, (d, score, label) in enumerate(versions):
        ws2.cell(row=start_row + 1 + i, column=start_col, value=d).font = BODY_FONT
        ws2.cell(row=start_row + 1 + i, column=start_col + 1, value=score).font = BODY_FONT
    last = start_row + len(versions)

    data_ref = Reference(ws2, min_col=start_col + 1, min_row=start_row, max_row=last)
    cats_ref = Reference(ws2, min_col=start_col, min_row=start_row + 1, max_row=last)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    s = chart.series[-1]
    s.graphicalProperties.line.width = 25000
    s.smooth = False

chart.y_axis.numFmt = '0.0'
chart.legend.position = "b"
ws2.add_chart(chart, f"A{best_start + 2 + len(best_samples) + 2}")

# ════════════════════════════════════════════
# SHEET 3 — MÉTRICAS (workflow performance)
# ════════════════════════════════════════════
ws3 = wb.create_sheet("Métricas")
ws3.cell(row=1, column=1, value="📈 Métricas Operacionais do Workflow").font = Font(name="Inter", bold=True, size=16, color="1A1A2E")
ws3.merge_cells("A1:H1")
ws3.row_dimensions[1].height = 40
ws3.cell(row=2, column=1, value="(Alimentado pelo step-log.ndjson — corre python3 scripts/log-metrics.py para atualizar)").font = Font(name="Inter", size=9, italic=True, color="888888")
ws3.merge_cells("A2:H2")

# Table 1: Per-Workflow Stats
ws3.cell(row=4, column=1, value="🏁 Performance por Workflow").font = Font(name="Inter", bold=True, size=12)
ws3.merge_cells("A4:G4")

wf_headers = ["Workflow ID", "Data", "Duração Total (s)", "Steps Total", "Sucesso", "Falhas", "Estado Final"]
for col, hdr in enumerate(wf_headers, 1):
    cell = ws3.cell(row=5, column=col, value=hdr)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER

ws3.column_dimensions["A"].width = 16
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 16
ws3.column_dimensions["D"].width = 14
ws3.column_dimensions["E"].width = 12
ws3.column_dimensions["F"].width = 12
ws3.column_dimensions["G"].width = 14

# Table 2: Per-Agent Reliability
agt_start = 8
ws3.cell(row=agt_start, column=1, value="🤖 Fiabilidade por Agente").font = Font(name="Inter", bold=True, size=12)
ws3.merge_cells(f"A{agt_start}:F{agt_start}")

agt_headers = ["Agente", "Modelo", "Steps", "Sucesso", "Taxa Sucesso", "Duração Média (s)"]
for col, hdr in enumerate(agt_headers, 1):
    cell = ws3.cell(row=agt_start + 1, column=col, value=hdr)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

# Table 3: Failure Trend
fail_start = agt_start + 9
ws3.cell(row=fail_start, column=1, value="❌ Análise de Falhas (últimas)").font = Font(name="Inter", bold=True, size=12)
ws3.merge_cells(f"A{fail_start}:G{fail_start}")

fail_headers = ["Data", "Workflow", "Step", "Agente", "Duração (ms)", "Erro"]
for col, hdr in enumerate(fail_headers, 1):
    cell = ws3.cell(row=fail_start + 1, column=col, value=hdr)
    cell.font = HEADER_FONT
    cell.fill = PatternFill(patternType="solid", fgColor="B71C1C")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

ws3.cell(row=fail_start + 2, column=1, value="(sem falhas registadas)").font = Font(name="Inter", italic=True, size=10, color="888888")
ws3.merge_cells(f"A{fail_start + 2}:F{fail_start + 2}")

# Table 4: Gates Status
gate_start = fail_start + 4
ws3.cell(row=gate_start, column=1, value="🚦 Estado dos Quality Gates").font = Font(name="Inter", bold=True, size=12)
ws3.merge_cells(f"A{gate_start}:E{gate_start}")

gate_headers = ["Gate", "Descrição", "Estado", "Último Workflow", "Observações"]
for col, hdr in enumerate(gate_headers, 1):
    cell = ws3.cell(row=gate_start + 1, column=col, value=hdr)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

gates = [
    ["G1: Wiki Knowledge", "wiki-keeper START", "✅", "TEST-001", "Sempre non-blocking"],
    ["G2: Plan Exists", "Plano válido em disco", "✅", "TEST-001", "Gate 3-active"],
    ["G3: Plan Approved", "review-plan + humano", "✅", "TEST-001", "Human approval gate"],
    ["G4: Code Coherent", "coherence-checker", "⏳", "—", "Ainda sem execuções"],
    ["G5: Code Quality", "SonarQube + princípios", "⏳", "—", "Ainda sem execuções"],
    ["G6: E2E Passed", "e2e-runner", "⏳", "—", "Ainda sem execuções"],
    ["G7: Knowledge Saved", "wiki-keeper END", "✅", "TEST-001", "Sempre non-blocking"],
]
for i, gate_data in enumerate(gates):
    row = gate_start + 2 + i
    for col, val in enumerate(gate_data, 1):
        cell = ws3.cell(row=row, column=col, value=val)
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN_BORDER

ws3.column_dimensions["D"].width = 20
ws3.column_dimensions["E"].width = 36

# ════════════════════════════════════════════
# SHEET 4 — CATEGORIES (reference)
# ════════════════════════════════════════════
ws4 = wb.create_sheet("Categorias")
ws4_header_cols = [(1, "Categoria"), (2, "Descrição")]
for col, val in ws4_header_cols:
    cell = ws4.cell(row=1, column=col, value=val)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER

categories_desc = {
    "Mudança de Tom": "Alteração de tom (formal → casual, técnico → simples)",
    "Foco/Âmbito": "Redefinir escopo do que o agente deve/não deve fazer",
    "Restrição": "Adicionar regras de segurança ou limitação",
    "Modelo": "Troca de modelo LLM",
    "Temperatura": "Ajuste de criatividade/determinismo",
    "Contexto": "Alteração no contexto fornecido ao agente",
    "Persona": "Mudança de papel/persona do agente",
    "Formato": "Alteração de formato de output",
    "Estrutura": "Reorganização do prompt (steps, secções)",
    "Correção Bug": "Correção de comportamento indesejado",
    "Outro": "Outras alterações",
}
for i, (cat, desc) in enumerate(sorted(categories_desc.items()), 2):
    ws4.cell(row=i, column=1, value=cat).font = BODY_FONT
    ws4.cell(row=i, column=1).border = THIN_BORDER
    ws4.cell(row=i, column=2, value=desc).font = BODY_FONT
    ws4.cell(row=i, column=2).border = THIN_BORDER

ws4.column_dimensions["A"].width = 22
ws4.column_dimensions["B"].width = 60

# ── Save ──
wb.save(FILE)
print(f"✅ Dashboard criado: {FILE}")
